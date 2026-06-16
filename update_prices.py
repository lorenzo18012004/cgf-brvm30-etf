"""
Met à jour la feuille Cours_Close de BRVM_Consolidated_Kendall_final.xlsx
avec les Cours Ajusté de BOURSE_AFRIQUE_COMPLET.xlsx.
Approche : copier le fichier original, puis écraser uniquement la feuille Cours_Close.
"""
import pandas as pd, numpy as np, sys, io, shutil, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

BA_FILE   = 'BOURSE_AFRIQUE_COMPLET.xlsx'
MAIN_FILE = 'BRVM_Consolidated_Kendall_final.xlsx'
OUT_FILE  = 'BRVM_Consolidated_Kendall_updated.xlsx'
COURS_SHEET = '📈 Cours_Close'

# ── 1. Charger tous les cours ajustés de BA ───────────────────────────────
print("Chargement BOURSE_AFRIQUE_COMPLET.xlsx...")
xl_ba = pd.ExcelFile(BA_FILE)
ba_data = {}
for t in xl_ba.sheet_names:
    df = xl_ba.parse(t)
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').set_index('Date')
    if 'Cours Ajuste' in df.columns:
        ba_data[t] = df['Cours Ajuste'].dropna().astype(float)
print(f"  {len(ba_data)} titres chargés depuis BA")

# ── 2. Charger la feuille cours de l'Excel principal ─────────────────────
print(f"Chargement {MAIN_FILE}...")
xl_main = pd.ExcelFile(MAIN_FILE)
prices_old = xl_main.parse(COURS_SHEET, index_col=0, parse_dates=True)
prices_old.index = pd.to_datetime(prices_old.index)
prices_old = prices_old.sort_index().astype(float)
print(f"  Prix actuels: {prices_old.shape[0]} dates × {prices_old.shape[1]} tickers")

# ── 3. Mettre à jour les prix ─────────────────────────────────────────────
prices_new = prices_old.copy()
updated_tickers = []
missing = []

for ticker in prices_old.columns:
    if ticker not in ba_data:
        missing.append(ticker)
        continue
    ba_series = ba_data[ticker]
    old_series = prices_old[ticker]
    common = old_series.dropna().index.intersection(ba_series.index)
    if len(common) < 10:
        missing.append(ticker)
        continue
    ratio = (old_series.reindex(common) / ba_series.reindex(common)).mean()
    n_diff = (old_series.reindex(common) / ba_series.reindex(common) - 1).abs().gt(0.01).sum()
    new_vals = old_series.copy()
    for date in old_series.dropna().index:
        if date in ba_series.index:
            new_vals[date] = ba_series[date]
    prices_new[ticker] = new_vals
    if n_diff > 0:
        updated_tickers.append(f"  {ticker}: {n_diff} jours modifiés (ratio={ratio:.4f})")

print(f"\nTickers mis à jour (prix changés):")
for u in updated_tickers:
    print(u)
print(f"Tickers absents de BA: {missing}")

# ── 4. Copier l'original et écraser uniquement la feuille Cours_Close ────
print(f"\nCopie de {MAIN_FILE} → {OUT_FILE}...")
shutil.copy2(MAIN_FILE, OUT_FILE)

print(f"Mise à jour de la feuille '{COURS_SHEET}'...")
wb = load_workbook(OUT_FILE)
if COURS_SHEET in wb.sheetnames:
    # Supprimer l'ancien contenu et recréer
    del wb[COURS_SHEET]

ws = wb.create_sheet(COURS_SHEET)

# Écrire les données ligne par ligne
# Header
ws.append(['Date'] + list(prices_new.columns))
# Data rows
for date_idx, row in prices_new.iterrows():
    row_data = [date_idx.date()] + [
        float(v) if pd.notna(v) else None for v in row.values
    ]
    ws.append(row_data)

# Replacer la feuille à sa position originale (1ère ou 2ème après README)
# Trouver la position originale
sheet_names = wb.sheetnames
if COURS_SHEET in sheet_names:
    target_idx = 1  # Après README
    current_idx = sheet_names.index(COURS_SHEET)
    if current_idx != target_idx:
        wb.move_sheet(COURS_SHEET, offset=target_idx - current_idx)

wb.save(OUT_FILE)
print(f"\nSauvegardé: {OUT_FILE}")
print(f"  {prices_new.shape[0]} dates × {prices_new.shape[1]} tickers")

# ── 5. Vérification rapide ────────────────────────────────────────────────
print("\nVérification...")
xl_check = pd.ExcelFile(OUT_FILE)
p_check = xl_check.parse(COURS_SHEET, index_col=0, parse_dates=True)
p_check.index = pd.to_datetime(p_check.index)
print(f"  Feuilles: {xl_check.sheet_names}")
print(f"  Cours_Close: {p_check.shape[0]} dates × {p_check.shape[1]} tickers")
if 'SAFC' in p_check.columns and 'SAFC' in prices_old.columns:
    old_safc = prices_old['SAFC'].loc['2023-01-01':'2023-03-31'].dropna()
    new_safc = p_check['SAFC'].loc['2023-01-01':'2023-03-31'].dropna()
    if len(old_safc) > 0 and len(new_safc) > 0:
        print(f"  SAFC jan2023 ancien={old_safc.iloc[0]:.0f} → nouveau={new_safc.iloc[0]:.0f}  (ratio={new_safc.iloc[0]/old_safc.iloc[0]:.4f})")

# Vérifier que BRVM_Indices est toujours lisible
try:
    idx_check = xl_check.parse('🏛️ BRVM_Indices', index_col=0, parse_dates=True)
    idx_check.index = pd.to_datetime(idx_check.index)
    brvm30_check = idx_check['BRVM30'].dropna()
    print(f"  BRVM_Indices OK: {len(brvm30_check)} pts BRVM30")
except Exception as e:
    print(f"  BRVM_Indices ERREUR: {e}")

print("\nDONE.")
