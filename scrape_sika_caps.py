"""
Scrape les données de capitalisation depuis Sika Finance (sikafinance.com)
pour tous les tickers BRVM et met à jour la feuille 🏢 Capitalisations de l'Excel.

Usage: python scrape_sika_caps.py
"""

import os, re, time, warnings
import requests
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd

warnings.filterwarnings('ignore')

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL     = os.path.join(BASE_DIR, 'BRVM_Consolidated_Kendall_updated.xlsx')
BASE_URL  = 'https://www.sikafinance.com/marches/societe/{ticker}.{pays}'

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/120.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'fr-FR,fr;q=0.9',
}

# Suffixes pays BRVM (déterminés par le siège social de l'émetteur)
COUNTRY_MAP = {
    'ABJC': 'ci', 'BICB': 'bj', 'BICC': 'ci', 'BNBC': 'ci',
    'BOAB': 'bj', 'BOABF': 'bf', 'BOAC': 'ci', 'BOAM': 'ml',
    'BOAN': 'ne', 'BOAS': 'sn', 'CABC': 'ci', 'CBIBF': 'bf',
    'CFAC': 'ci', 'CIEC': 'ci', 'ECOC': 'ci', 'ETIT': 'tg',
    'FTSC': 'ci', 'LNBB': 'bj', 'NEIC': 'ci', 'NSBC': 'ci',
    'NTLC': 'ci', 'ONTBF': 'bf', 'ORAC': 'ci', 'ORGT': 'tg',
    'PALC': 'ci', 'PRSC': 'ci', 'SAFC': 'ci', 'SCRC': 'ci',
    'SDCC': 'ci', 'SDSC': 'ci', 'SEMC': 'ci', 'SGBC': 'ci',
    'SHEC': 'ci', 'SIBC': 'ci', 'SICC': 'ci', 'SIVC': 'ci',
    'SLBC': 'ci', 'SMBC': 'ci', 'SNTS': 'sn', 'SOGC': 'ci',
    'SPHC': 'ci', 'STAC': 'ci', 'STBC': 'ci', 'TTLC': 'ci',
    'TTLS': 'sn', 'UNLC': 'ci', 'UNXC': 'ci',
}


def _clean_number(s: str) -> float:
    """Convertit '57 759 800' ou '309 015 MFCFA' en float."""
    s = re.sub(r'[^\d,\.]', '', s.replace(' ', '').replace('\xa0', '').replace(' ', ''))
    s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def scrape_ticker(ticker: str, pays: str) -> dict | None:
    """Retourne dict avec nb_total, flottant_pct, valorisation_mfcfa ou None."""
    url = BASE_URL.format(ticker=ticker, pays=pays)
    try:
        r = requests.get(url, headers=HEADERS, timeout=20, verify=False)
        if r.status_code != 200:
            # Essayer sans suffixe pays ou avec 'ci' par défaut
            if pays != 'ci':
                r2 = requests.get(
                    BASE_URL.format(ticker=ticker, pays='ci'),
                    headers=HEADERS, timeout=15, verify=False
                )
                if r2.status_code == 200:
                    r = r2
                else:
                    return None
            else:
                return None
        soup = BeautifulSoup(r.text, 'html.parser')
        text = soup.get_text()

        nb_total   = None
        flottant   = None
        valo_mfcfa = None

        for line in text.split('\n'):
            line = line.strip()
            # Nombre de titres
            m = re.search(r'Nombre de titres\s*:?\s*([\d\s \xa0]+)', line, re.I)
            if m:
                nb_total = _clean_number(m.group(1))
            # Flottant %
            m = re.search(r'Flottant\s*:?\s*([\d,\.]+)\s*%', line, re.I)
            if m:
                flottant = float(m.group(1).replace(',', '.')) / 100
            # Valorisation
            m = re.search(r'Valorisation.*?([\d\s \xa0]+)\s*MFCFA', line, re.I)
            if m:
                valo_mfcfa = _clean_number(m.group(1))

        if nb_total and flottant is not None:
            return {
                'nb_total':     int(nb_total),
                'flottant_pct': flottant,
                'valo_mfcfa':   valo_mfcfa,
            }
        return None

    except Exception as e:
        print(f"  [ERR] {ticker}: {e}")
        return None


def main():
    wb   = openpyxl.load_workbook(EXCEL)
    ws   = wb['🏢 Capitalisations']

    header   = [str(c.value).strip() if c.value else '' for c in ws[1]]
    col_code = header.index('Code') + 1
    col_capfl = header.index('Cap. Flottante (FCFA)') + 1
    col_capgl = header.index('Cap. Globale (FCFA)') + 1
    col_ratio = header.index('Ratio Flottant/Global') + 1
    col_nbfl  = header.index('Nb Titres Flottants') + 1
    col_nbtot = header.index('Nb Titres Total') + 1

    # Index des lignes par code
    row_map = {}
    for row in ws.iter_rows(min_row=2):
        code = str(row[col_code-1].value).strip() if row[col_code-1].value else ''
        if code:
            row_map[code] = row

    updated  = 0
    failed   = 0
    skipped  = 0
    results  = {}

    tickers = list(COUNTRY_MAP.keys())
    print(f"Scraping {len(tickers)} tickers sur Sika Finance...")
    print()

    for ticker in tickers:
        pays = COUNTRY_MAP.get(ticker, 'ci')
        data = scrape_ticker(ticker, pays)
        time.sleep(0.5)  # polir politesse

        if data is None:
            print(f"  ✗ {ticker:6s} — non trouvé")
            failed += 1
            continue

        results[ticker] = data
        nb_total = data['nb_total']
        ratio    = data['flottant_pct']
        nb_flott = round(nb_total * ratio)

        if data['valo_mfcfa']:
            cap_gl = int(data['valo_mfcfa'] * 1_000_000)
        else:
            cap_gl = None

        cap_fl = round(cap_gl * ratio) if cap_gl else None

        if ticker not in row_map:
            print(f"  ? {ticker:6s} — absent de la feuille Excel")
            skipped += 1
            continue

        row = row_map[ticker]
        if cap_gl:
            row[col_capgl-1].value = cap_gl
        if cap_fl:
            row[col_capfl-1].value = cap_fl
        row[col_ratio-1].value = round(ratio, 4)
        row[col_nbfl-1].value  = nb_flott
        row[col_nbtot-1].value = nb_total
        updated += 1

        cap_str = f"{cap_fl/1e9:.1f} Mds" if cap_fl else "N/A"
        print(f"  ✓ {ticker:6s}  flott={ratio*100:.0f}%  nb_titres={nb_total:,}  cap_fl={cap_str}")

    print()
    print(f"Résultats: {updated} mis à jour, {failed} non trouvés, {skipped} absents de l'Excel")

    wb.save(EXCEL)
    print(f"Excel sauvegardé → {EXCEL}")
    return results


if __name__ == '__main__':
    main()
