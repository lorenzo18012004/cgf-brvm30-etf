"""
scrape_sika.py — Mise à jour quotidienne des cours depuis Sika Finance
=======================================================================
Récupère les cours de clôture du jour sur sikafinance.com/marches/aaz
et les ajoute à BRVM_Consolidated_Kendall_updated.xlsx (feuille Cours_Close).

Usage :
    python scrape_sika.py              # mise à jour normale + affichage
    python scrape_sika.py --dry-run    # simulation sans écriture
    python scrape_sika.py --date 2026-05-23   # forcer une date précise

Puis enchaîner avec :
    python calc_nav.py
"""

import sys, io, os, re, json, argparse, shutil, warnings
warnings.filterwarnings('ignore')

import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from base import BaseScript


class SikaScraper(BaseScript):

    def __init__(self):
        super().__init__()
        self.sika_url         = 'https://sikafinance.com/marches/aaz'
        self.prix_sheet       = '📈 Cours_Close'
        self.headers          = {
            'User-Agent': (
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                'AppleWebKit/537.36 (KHTML, like Gecko) '
                'Chrome/124.0.0.0 Safari/537.36'
            ),
            'Accept-Language': 'fr-FR,fr;q=0.9,en;q=0.8',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        }
        self.brvm30_hist_url  = 'https://www.sikafinance.com/marches/historiques/BRVM30'
        self.brvm30_hist_file = os.path.join(self.data_dir, 'brvm30_index_history.json')
        self.richbourse_history = os.path.join(self.data_dir, 'richbourse_history.json')

    def _fetch_html(self, url, timeout = 20):
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        resp = requests.get(url, headers=self.headers, timeout=timeout, verify=False)
        resp.raise_for_status()
        return resp.text

    def _ticker_from_href(self, href):
        """Extrait le ticker depuis un href type '/marches/cotation_SNTS.ci'."""
        m = re.search(r'cotation_([A-Z0-9]+)', href, re.IGNORECASE)
        if not m:
            return None
        code = m.group(1).upper()
        return code

    def scrape_brvm30_history(self, max_rows = 500):
        """
        Scrape la page historiques/BRVM30 et retourne {date_iso: close_value}.
        Date format sur Sika : DD/MM/YYYY → converti en YYYY-MM-DD.
        """
        html = self._fetch_html(self.brvm30_hist_url)
        soup = BeautifulSoup(html, 'html.parser')
        result = {}

        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            for row in rows[1:max_rows]:
                cells = [td.get_text(strip=True).replace('\xa0', '').replace(' ', '') for td in row.find_all(['td', 'th'])]
                if len(cells) < 2:
                    continue
                date_txt = cells[0]
                m = re.match(r'(\d{2})/(\d{2})/(\d{4})', date_txt)
                if not m:
                    continue
                date_iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
                close_txt = cells[1].replace(',', '.')
                try:
                    close_val = float(close_txt)
                    if 50 < close_val < 5000:
                        result[date_iso] = round(close_val, 4)
                except ValueError:
                    continue
            if result:
                break

        return result

    def save_brvm30_index(self, value, trade_date):
        """Sauvegarde/met à jour une valeur de l'indice BRVM30 dans brvm30_index_history.json."""
        hist = {}
        if os.path.exists(self.brvm30_hist_file):
            with open(self.brvm30_hist_file, encoding='utf-8') as f:
                try:
                    hist = json.load(f)
                except Exception:
                    hist = {}
        hist[str(trade_date)] = round(value, 4)
        with open(self.brvm30_hist_file, 'w', encoding='utf-8') as f:
            json.dump(hist, f, ensure_ascii=False, indent=2)

    def scrape_brvm30_index(self, html):
        """
        Extrait la valeur courante de l'indice BRVM30 depuis la page aaz.
        Retourne None si non trouvé.
        """
        soup = BeautifulSoup(html, 'html.parser')
        for a in soup.find_all('a', href=re.compile(r'/marches/cotation_BRVM30', re.I)):
            row = a.find_parent('tr')
            if not row:
                continue
            cells = row.find_all(['td', 'th'])
            for cell in reversed(cells):
                txt = cell.get_text(strip=True).replace('\xa0', '').replace(' ', '').replace(',', '.')
                try:
                    val = float(txt)
                    if 50 < val < 5000:
                        return val
                except ValueError:
                    continue
        return None

    def scrape_prices(self, html):
        """
        Parse la page aaz et renvoie une Series {ticker: prix_cloture}.
        Stratégie 1 : BeautifulSoup (extrait ticker via href + prix via colonnes)
        Stratégie 2 : pandas read_html (fallback)
        """
        soup = BeautifulSoup(html, 'html.parser')

        prices = {}
        for a in soup.find_all('a', href=re.compile(r'/marches/cotation_[A-Z]', re.I)):
            ticker = self._ticker_from_href(a['href'])
            if not ticker:
                continue
            if any(x in ticker for x in ('BRVM', 'SIKA', 'COMPO')):
                continue

            row = a.find_parent('tr')
            if not row:
                continue
            cells = row.find_all(['td', 'th'])
            if len(cells) < 4:
                continue

            raw_price = None
            for cell in reversed(cells):
                txt = cell.get_text(strip=True).replace('\xa0', '').replace(' ', '')
                txt = txt.replace(',', '.').replace(' ', '')
                try:
                    val = float(txt)
                    if 10 < val < 1_000_000:
                        raw_price = val
                        break
                except ValueError:
                    continue

            if raw_price is not None:
                prices[ticker] = raw_price

        if len(prices) >= 10:
            print(f"  Strategie 1 (BeautifulSoup) : {len(prices)} tickers extraits")
            return pd.Series(prices)

        print("  Strategie 1 insuffisante, tentative avec pandas read_html...")
        try:
            tables = pd.read_html(html)
            target = None
            for t in tables:
                cols_lower = [str(c).lower() for c in t.columns]
                if any('dernier' in c or 'cours' in c for c in cols_lower):
                    target = t
                    break
            if target is None and tables:
                target = max(tables, key=len)

            if target is not None:
                col_prix = None
                for c in target.columns:
                    if 'dernier' in str(c).lower() or 'cours' in str(c).lower():
                        col_prix = c
                        break
                if col_prix is None:
                    col_prix = target.columns[-2]

                prices2 = {}
                for _, row in target.iterrows():
                    nom = str(row.iloc[0])
                    m = re.match(r'^([A-Z]{3,6})', nom)
                    if m:
                        tk = m.group(1)
                        try:
                            val = float(str(row[col_prix]).replace(',', '.').replace(' ', ''))
                            if 10 < val < 1_000_000:
                                prices2[tk] = val
                        except (ValueError, KeyError):
                            pass
                if prices2:
                    print(f"  Strategie 2 (read_html) : {len(prices2)} tickers extraits")
                    return pd.Series(prices2)

        except Exception as e:
            print(f"  Strategie 2 echouee : {e}")

        raise RuntimeError(
            "Impossible de parser les cours depuis sikafinance.com/marches/aaz.\n"
            "La structure du site a peut-être changé. Vérifiez manuellement."
        )

    def _load_existing_prices(self):
        xl = pd.ExcelFile(self.prix_file)
        prices = xl.parse(self.prix_sheet, index_col=0, parse_dates=True)
        prices.index = pd.to_datetime(prices.index)
        prices = prices.sort_index().astype(float)
        return prices

    def _last_trading_date(self):
        """Dernier jour de bourse (lun–ven, heure Abidjan = UTC).
        Utilise aujourd'hui si le marché est ouvert (09h-18h UTC), sinon hier.
        """
        today = datetime.utcnow().date()
        while today.weekday() >= 5:
            today -= timedelta(days=1)
        return today

    def update_excel(
        self,
        scraped: pd.Series,
        trade_date: date,
        existing: pd.DataFrame,
        dry_run: bool = False,
    ):
        """
        Ajoute une nouvelle ligne de cours à l'Excel.
        Renvoie un dict de résumé.
        """
        ts = pd.Timestamp(trade_date)

        if ts in existing.index:
            return {
                'status': 'already_exists',
                'date': str(trade_date),
                'message': f"Les cours du {trade_date} sont deja dans l'Excel.",
            }

        new_row = existing.iloc[-1].copy() * np.nan
        matched, unmatched = [], []

        for ticker, price in scraped.items():
            if ticker in existing.columns:
                new_row[ticker] = price
                matched.append(ticker)
            else:
                unmatched.append(ticker)

        new_df = pd.concat([existing, pd.DataFrame([new_row], index=[ts])])
        new_df = new_df.sort_index()

        summary = {
            'status': 'ok',
            'date': str(trade_date),
            'n_matched': len(matched),
            'n_unmatched': len(unmatched),
            'n_total_scraped': len(scraped),
            'tickers_missing': unmatched[:10],
            'tickers_updated': matched,
        }

        if dry_run:
            summary['status'] = 'dry_run'
            return summary

        backup = self.prix_file.replace('.xlsx', f'_backup_{trade_date}.xlsx')
        shutil.copy2(self.prix_file, backup)

        wb = load_workbook(self.prix_file)
        if self.prix_sheet in wb.sheetnames:
            del wb[self.prix_sheet]

        ws = wb.create_sheet(self.prix_sheet)
        ws.append(['Date'] + list(new_df.columns))
        for idx, row in new_df.iterrows():
            row_data = [idx.date()] + [
                float(v) if pd.notna(v) else None for v in row.values
            ]
            ws.append(row_data)

        sheet_names = wb.sheetnames
        if self.prix_sheet in sheet_names:
            current_idx = sheet_names.index(self.prix_sheet)
            if current_idx != 1:
                wb.move_sheet(self.prix_sheet, offset=1 - current_idx)

        wb.save(self.prix_file)

        try:
            os.remove(backup)
        except OSError:
            pass

        return summary

    def _print_summary(self, s, scraped):
        W = 60
        print("=" * W)
        print("  Sika Finance → BRVM Consolidated")
        print("=" * W)
        print(f"  Date de cotation   : {s['date']}")
        if 'n_total_scraped' in s:
            print(f"  Tickers scrapes    : {s['n_total_scraped']}")
        if 'n_matched' in s:
            print(f"  Matches Excel      : {s['n_matched']}")

        if s['status'] == 'already_exists':
            print(f"\n  [OK] {s['message']}")
        elif s['status'] == 'dry_run':
            print(f"\n  [DRY-RUN] Aucune ecriture effectuee.")
            print(f"  Exemples de cours scraped :")
            for t, p in list(scraped.items())[:8]:
                print(f"    {t:<8} : {p:>8,.0f} FCFA")
        else:
            print(f"\n  [OK] Excel mis a jour.")
            print(f"  Relancer : python calc_nav.py")

        if s.get('tickers_missing'):
            print(f"\n  Tickers Sika non reconnus : {s['tickers_missing']}")
        print("=" * W)

    def _fallback_from_richbourse(self, trade_date):
        """Retourne les derniers cours connus depuis richbourse_history.json."""
        if not os.path.exists(self.richbourse_history):
            raise FileNotFoundError("richbourse_history.json introuvable — pas de fallback possible.")
        with open(self.richbourse_history, encoding='utf-8') as f:
            hist = json.load(f)
        date_str = str(trade_date)
        prices = {}
        for ticker, dates_dict in hist.items():
            past = sorted(d for d in dates_dict if d <= date_str)
            if past:
                prices[ticker.upper()] = float(dates_dict[past[-1]])
        if not prices:
            raise ValueError("richbourse_history.json ne contient aucun cours utilisable.")
        return pd.Series(prices)

    def run(self, trade_date = None, dry_run = False):
        os.chdir(self.data_dir)

        if trade_date is None:
            trade_date = self._last_trading_date()

        print(f"Scraping sikafinance.com pour le {trade_date}...")

        scraped = None
        source = 'sika'
        try:
            html    = self._fetch_html(self.sika_url)
            scraped = self.scrape_prices(html)
        except Exception as e:
            print(f"  [WARN] Sika Finance indisponible : {e}")
            print("  Tentative fallback Richbourse...")
            try:
                scraped = self._fallback_from_richbourse(trade_date)
                source  = 'richbourse_fallback'
                print(f"  [FALLBACK] {len(scraped)} cours chargés depuis richbourse_history.json")
            except Exception as e2:
                print(f"  [ERREUR] Fallback échoué : {e2}")
                return {'status': 'network_error', 'error': str(e), 'fallback_error': str(e2)}
        print(f"  {len(scraped)} cours recuperes (source: {source})")

        existing = self._load_existing_prices()
        print(f"  Excel actuel : {existing.shape[0]} dates x {existing.shape[1]} tickers")

        if source == 'sika':
            try:
                hist_data = self.scrape_brvm30_history()
                if hist_data:
                    existing_hist = {}
                    if os.path.exists(self.brvm30_hist_file):
                        with open(self.brvm30_hist_file, encoding='utf-8') as _f:
                            try:
                                existing_hist = json.load(_f)
                            except Exception:
                                existing_hist = {}
                    existing_hist.update(hist_data)
                    with open(self.brvm30_hist_file, 'w', encoding='utf-8') as _f:
                        json.dump(existing_hist, _f, ensure_ascii=False, indent=2)
                    today_val = hist_data.get(str(trade_date))
                    print(f"  BRVM30 historique : {len(hist_data)} dates récupérées"
                          + (f" — aujourd'hui : {today_val}" if today_val else ""))
            except Exception as _e:
                print(f"  [WARN] BRVM30 historique non récupéré : {_e}")

        summary = self.update_excel(scraped, trade_date, existing, dry_run=dry_run)
        summary['source'] = source
        self._print_summary(summary, scraped)
        return summary


# Aliases module-level pour les scripts qui importent ces fonctions directement
SIKA_URL = 'https://sikafinance.com/marches/aaz'

def _fetch_html(url, timeout = 20):
    return SikaScraper()._fetch_html(url, timeout)

def scrape_prices(html):
    return SikaScraper().scrape_prices(html)

def scrape_brvm30_index(html):
    return SikaScraper().scrape_brvm30_index(html)


if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    parser = argparse.ArgumentParser(description='Scraper cours Sika Finance → Excel BRVM')
    parser.add_argument('--dry-run', action='store_true',
                        help='Simulation sans modification de l\'Excel')
    parser.add_argument('--date', type=str, default=None,
                        help='Forcer une date de cotation (format YYYY-MM-DD)')
    args = parser.parse_args()

    td = None
    if args.date:
        td = date.fromisoformat(args.date)

    SikaScraper().run(trade_date=td, dry_run=args.dry_run)
