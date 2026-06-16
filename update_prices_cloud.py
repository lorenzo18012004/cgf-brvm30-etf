"""
update_prices_cloud.py — Mise à jour quotidienne des prix (version cloud)
=========================================================================
Scrape les cours de clôture depuis Sika Finance et met à jour :
  - richbourse_history.json
  - BRVM_Consolidated_Kendall_updated.xlsx (feuille Cours_Close)

Compatible GitHub Actions — remplace update_daily.py pour le cloud.
Usage : python update_prices_cloud.py [--date YYYY-MM-DD]
"""
import os, sys, json, re, argparse, warnings
warnings.filterwarnings('ignore')

import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import date, datetime

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
HIST_FILE  = os.path.join(BASE_DIR, 'richbourse_history.json')
EXCEL_FILE = os.path.join(BASE_DIR, 'BRVM_Consolidated_Kendall_updated.xlsx')
SHEET      = '📈 Cours_Close'
SIKA_URL   = 'https://sikafinance.com/marches/aaz'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
    'Accept-Language': 'fr-FR,fr;q=0.9',
}


def scrape_closing_prices() -> dict:
    """Scrape tous les cours de clôture depuis Sika Finance."""
    resp = requests.get(SIKA_URL, headers=HEADERS, verify=False, timeout=20)
    resp.raise_for_status()
    soup    = BeautifulSoup(resp.text, 'html.parser')
    prices  = {}

    for a in soup.find_all('a', href=re.compile(r'/marches/cotation_', re.I)):
        m = re.search(r'cotation_([A-Z0-9]+)', a['href'], re.I)
        if not m:
            continue
        ticker = m.group(1).upper()
        if any(x in ticker for x in ('BRVM', 'SIKA', 'COMPO')):
            continue
        row = a.find_parent('tr')
        if not row:
            continue
        cells = [td.get_text(strip=True) for td in row.find_all(['td', 'th'])]
        for cell in cells:
            val = cell.replace('\xa0', '').replace(' ', '').replace(',', '.')
            try:
                px = float(val)
                if px > 0:
                    prices[ticker] = px
                    break
            except ValueError:
                continue

    print(f"[Sika] {len(prices)} cours récupérés")
    return prices


def update_history(prices: dict, date_str: str) -> None:
    """Met à jour richbourse_history.json avec les prix du jour."""
    hist = {}
    if os.path.exists(HIST_FILE):
        with open(HIST_FILE, encoding='utf-8') as f:
            hist = json.load(f)

    for ticker, px in prices.items():
        if ticker not in hist:
            hist[ticker] = {}
        hist[ticker][date_str] = px

    with open(HIST_FILE, 'w', encoding='utf-8') as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)
    print(f"[richbourse_history.json] {len(prices)} prix ajoutés pour {date_str}")


def update_excel(prices: dict, date_str: str) -> None:
    """Met à jour l'Excel BRVM_Consolidated_Kendall_updated.xlsx."""
    if not os.path.exists(EXCEL_FILE):
        print(f"[WARN] Excel introuvable : {EXCEL_FILE}")
        return

    xl      = pd.ExcelFile(EXCEL_FILE)
    df      = xl.parse(SHEET, index_col=0, parse_dates=True)
    df.index = pd.to_datetime(df.index)
    df      = df.sort_index().astype(float)

    ts = pd.Timestamp(date_str)
    if ts in df.index:
        print(f"[Excel] {date_str} déjà présent — mise à jour")
        for ticker, px in prices.items():
            if ticker in df.columns:
                df.loc[ts, ticker] = px
    else:
        new_row = {col: prices.get(col, float('nan')) for col in df.columns}
        df = pd.concat([df, pd.DataFrame([new_row], index=[ts])])

    df = df.sort_index()

    # Réécrire la feuille
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = load_workbook(EXCEL_FILE)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    # En-tête
    ws.append(['Date'] + list(df.columns))
    for idx, row in df.iterrows():
        ws.append([idx.strftime('%Y-%m-%d')] + [v if not pd.isna(v) else None for v in row])

    wb.save(EXCEL_FILE)
    print(f"[Excel] {date_str} ecrit ({len(df)} lignes)")


def run(date_str: str | None = None) -> None:
    if date_str is None:
        date_str = date.today().strftime('%Y-%m-%d')

    print(f"\n=== Mise à jour des prix cloud — {date_str} ===")
    prices = scrape_closing_prices()

    if not prices:
        print("[ERREUR] Aucun cours récupéré")
        sys.exit(1)

    update_history(prices, date_str)
    update_excel(prices, date_str)
    print("=== Terminé ===\n")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--date', default=None)
    args = parser.parse_args()
    run(args.date)
